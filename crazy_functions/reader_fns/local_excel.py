# encoding: utf-8
# @Time   : 2024/1/15
# @Author : Spike
# @Descr   : Excel 文件处理工具
import os

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Font

from crazy_functions.reader_fns import crazy_box
from common import func_box
from common import toolbox

import xmind


class XlsxHandler:

    def __init__(self, xlsx_path, output_dir=None, sheet='测试要点'):
        """
        Args:
            xlsx_path: 文件路径
            output_dir: 保存路径，如果没有另存为操作，可以为空
            sheet:
        """
        if output_dir:
            self.output_dir = os.path.join(output_dir, 'excel')
            os.makedirs(self.output_dir, exist_ok=True)
        self.template_excel = xlsx_path
        self.sheet = sheet
        self.workbook = load_workbook(self.template_excel)
        self.file_name = os.path.splitext(os.path.basename(self.template_excel))[0]
        # 定义填充样式
        self.yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        self.green_fill = PatternFill(start_color="1abc9c", end_color="1abc9c", fill_type="solid")
        self.red_fill = PatternFill(start_color="ff7f50", end_color="ff7f50", fill_type="solid")
        # 定义边框样式
        border_style = Side(style='thin', color="000000")
        self.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
        if str(self.sheet) not in self.workbook.sheetnames:
            self.sheet = self.workbook.active.title

    def list_write_to_excel(self, data_list: list, save_as_name=''):
        # 加载现有的 Excel 文件
        if self.sheet in self.workbook.sheetnames:
            worksheet = self.workbook[self.sheet]
        else:
            worksheet = self.workbook.create_sheet(self.sheet)
        # 定义起始行号
        start_row = crazy_box.find_index_inlist(self.read_as_dict()[self.sheet],
                                                ['前置条件', '操作步骤', '预期结果']) + 2
        # 创建一个黄色的填充样式
        # 遍历数据列表
        for row_data in data_list:
            # 写入每一行的数据到指定的单元格范围
            for col_num, value in enumerate(row_data, start=1):
                cell = worksheet.cell(row=start_row, column=col_num)
                try:
                    cell.value = str(value).strip()
                    cell.border = self.border
                    cell.alignment = Alignment(horizontal='left', vertical='center',
                                               wrapText=True)  # 设置水平方向居左对齐，并垂直方向居中对齐，并自动换行
                    # 判断 value 是否为 '插件补充的用例'
                    if '插件补充的用例' in str(value):
                        cell.fill = self.yellow_fill
                    font = Font(name='苹方-简', size=11)
                    cell.font = font
                except Exception:
                    print(row_data, value)
                    func_box.通知机器人(error=f'写入excel错误啦\n\n```\n\n{row_data}\n\n{value}\n\n```'
                                              f'\n\n```\n\n{toolbox.trimmed_format_exc()}```\n\n')
            # 增加起始行号
            start_row += 1
        merge_cell = toolbox.get_conf('merge_cell')
        if merge_cell:
            self.merge_same_cells()  # 还原被拆分的合并单元格
        if save_as_name:
            save_as_name = f'{save_as_name}_'
        test_case_path = os.path.join(self.output_dir, f'{save_as_name}.xlsx')
        # 遇到文件无法保存时，再拆开图片
        try:
            self.workbook.save(test_case_path)
        except Exception as f:
            test_case_path = self.template_excel
        return test_case_path

    def read_as_dict(self, only_sheet=True):
        data_dict = {}
        # 遍历每个工作表
        if only_sheet:
            sheet_list = [self.sheet]
        else:
            sheet_list = self.workbook.sheetnames
        for sheet_name in sheet_list:
            sheet = self.workbook[sheet_name]
            sheet_data = []
            # 遍历每一行
            sheet_temp_count = 0
            for row in sheet.iter_rows(values_only=True):
                # 过滤尾部的空行
                row = tuple(x for x in row if x is not None and x != row[-1])
                if row:
                    sheet_data.append(row)
                else:
                    sheet_temp_count += 1
                if sheet_temp_count >= 20: break
            # 将工作表名作为字典的键，行数据作为值
            data_dict[sheet_name] = sheet_data
        return data_dict

    def split_merged_cells(self):
        # 加载Excel文件
        ws = self.workbook[self.sheet]
        # 获取合并单元格的范围
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            # 获取合并单元格的起始行、起始列、结束行、结束列
            start_row = merged_range.min_row
            start_col = merged_range.min_col
            end_row = merged_range.max_row
            end_col = merged_range.max_col
            # 获取合并单元格的值
            value = ws.cell(start_row, start_col).value
            # 拆分合并单元格
            ws.unmerge_cells(str(merged_range))
            # 在每个拆分后的单元格中填入值
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row, col)
                    cell.value = value
        # 保存结果
        self.workbook.save(self.template_excel)

    def merge_same_cells(self, truncation=10):
        # 加载xlsx文件
        ws = self.workbook[self.sheet]
        # 遍历每个单元格（列优先遍历）
        column_counter = {'row': 0, 'col': 0}
        for col_index in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_index)
            row_start = None
            last_column_empty = True
            for row_index in range(1, ws.max_row + 1):
                current_cell = ws[f"{col_letter}{row_index}"]
                next_cell = ws[f"{col_letter}{row_index + 1}"]
                # 当前单元格与下个单元格内容相同时，都不为空，并记录合并范围row_start
                if row_start is None and current_cell.value == next_cell.value and current_cell.value is not None:
                    row_start = row_index
                # 当前单元格与下个单元格内容不同时或任何一个为空时，记录合并范围row_end，并执行合并
                elif row_start is not None and (
                        current_cell.value != next_cell.value or current_cell.value is None or next_cell.value is None):
                    row_end = row_index
                    ws.merge_cells(f"{col_letter}{row_start}:{col_letter}{row_end}")
                    row_start = None
                # # 设置边框样式
                current_cell.border = self.border
                next_cell.border = self.border
                # 当列超过10行为空，跳出循环
                if not current_cell.value:
                    column_counter['row'] += 1
                    if column_counter['row'] > truncation:
                        column_counter['row'] = 0
                        break
                # 检查当前列是否为空
            if all(cell.value is None for cell in ws[col_letter]):
                if last_column_empty:  # 如果上一列也为空，增加计数器
                    column_counter['col'] += 1
                    if column_counter['col'] > truncation:  # 如果空列超过所设定的上限，跳出循环
                        break
                else:  # 如果上一列非空，重置计数器
                    column_counter['col'] = 1
                    last_column_empty = True
            else:  # 如果当前列非空，重置计数器和 last_column_empty 标记
                last_column_empty = False
                column_counter['col'] = 0
        self.workbook.save(self.template_excel)

    def to_xmind(self):
        """只能用 xmind8 打开 心态炸了"""
        file_path = os.path.join(self.output_dir, f"{self.file_name}.xmind")
        # 创建一个新的XMind工作簿
        xmind_book = xmind.load(file_path)  # 创建新的工作簿
        fields_dict = self.read_as_dict()
        for topic in fields_dict:
            x_sheet = xmind_book.createSheet()
            x_sheet.setTitle(topic)
            root_topic = x_sheet.getRootTopic()  # 获取根主题
            root_topic.setTitle(topic)  # 设置根主题标题为用例标签
            for data in fields_dict[topic]:
                for i, sub_data in enumerate(data):
                    child_topic = root_topic.addSubTopic()  # 创建子子主题
                    child_topic.setTitle(sub_data)
                    root_topic.addSubTopic(child_topic)  # 将子主题添加到根主题下
        # 保存工作簿
        xmind.save(xmind_book, file_path)


if __name__ == '__main__':
    pass
