#! .\venv\
# encoding: utf-8
# @Time   : 2023/6/14
# @Author : Spike
# @Descr   :
import os
import json
import requests
import re
import time
from comm_tools import func_box, ocr_tools, toolbox, prompt_generator
from openpyxl import load_workbook
from crazy_functions import crazy_utils
from request_llm import bridge_all
from crazy_functions import crzay_kingsoft


class Utils:

    def __init__(self):
        self.find_keys_type = 'type'
        self.find_picture_source = ['caption', 'imgID', 'sourceKey']
        self.find_document_source = ['wpsDocumentLink', 'wpsDocumentName', 'wpsDocumentType']
        self.find_document_tags = ['WPSDocument']
        self.find_picture_tags = ['picture', 'processon']
        self.picture_format = ['.jpeg', '.jpg', '.png', '.gif', '.bmp', '.tiff']

    def find_all_text_keys(self, dictionary, parent_type=None, text_values=None, filter_type=''):
        """
        Args:
            dictionary: 字典或列表
            parent_type: 匹配的type，作为新列表的key，用于分类
            text_values: 存储列表
            filter_type: 当前层级find_keys_type==filter_type时，不继续往下嵌套
        Returns:
            text_values和排序后的context_
        """
        # 初始化 text_values 为空列表，用于存储找到的所有text值
        if text_values is None:
            text_values = []
        # 如果输入的dictionary不是字典类型，返回已收集到的text值
        if not isinstance(dictionary, dict):
            return text_values
        # 获取当前层级的 type 值
        current_type = dictionary.get(self.find_keys_type, parent_type)
        # 如果字典中包含 'text' 或 'caption' 键，将对应的值添加到 text_values 列表中
        if 'text' in dictionary:
            content_value = dictionary.get('text', None)
            text_values.append({current_type: content_value})
        if 'caption' in dictionary:
            temp = {}
            for key in self.find_picture_source:
                temp[key] = dictionary.get(key, None)
            text_values.append({current_type: temp})
        if 'wpsDocumentId' in dictionary:
            temp = {}
            for key in self.find_document_source:
                temp[key] = dictionary.get(key, None)
            text_values.append({current_type: temp})
        # 如果当前类型不等于 filter_type，则继续遍历子级属性
        if current_type != filter_type:
            for key, value in dictionary.items():
                if isinstance(value, dict):
                    self.find_all_text_keys(value, current_type, text_values, filter_type)
                elif isinstance(value, list):
                    for item in value:
                        if isinstance(item, dict):
                            self.find_all_text_keys(item, current_type, text_values, filter_type)
        return text_values

    def statistical_results(self, text_values, img_proce=False):
        """
        Args: 提取爬虫内嵌图片、文件等等信息
            text_values: dict
            img_proce: 图片标识
        Returns: （元数据， 组合数据， 图片dict， 文件dict）
        """
        context_ = []
        pic_dict = {}
        file_dict = {}
        for i in text_values:
            for key, value in i.items():
                if key in self.find_picture_tags:
                    if img_proce:
                        mark = f'{key}OCR结果: """{value["sourceKey"]}"""\n'
                        if value["caption"]: mark += f'{key}描述: {value["caption"]}\n'
                        context_.append(mark)
                        pic_dict[value['sourceKey']] = value['imgID']
                    else:
                        if value["caption"]: context_.append(f'{key}描述: {value["caption"]}\n')
                        pic_dict[value['sourceKey']] = value['imgID']
                elif key in self.find_document_tags:
                    mark = f'{value["wpsDocumentName"]}: {value["wpsDocumentLink"]}'
                    file_dict.update({value["wpsDocumentName"]: value["wpsDocumentLink"]})
                    context_.append(mark)
                else:
                    context_.append(value)
        context_ = '\n'.join(context_)
        return text_values, context_, pic_dict, file_dict

    def write_markdown(self, data, hosts, file_name):
        """
        Args: 将data写入md文件
            data: 数据
            hosts: 用户标识
            file_name: 另取文件名
        Returns: 写入的文件地址
        """
        user_path = os.path.join(func_box.users_path, hosts, 'markdown')
        os.makedirs(user_path, exist_ok=True)
        md_file = os.path.join(user_path, f"{file_name}.md")
        with open(file=md_file, mode='w') as f:
            f.write(data)
        return md_file

    def markdown_to_flow_chart(self, data, hosts, file_name):
        """
        Args: 调用markmap-cli
            data: 要写入md的数据
            hosts: 用户标识
            file_name: 要写入的文件名
        Returns: [md, 流程图] 文件
        """
        user_path = os.path.join(func_box.users_path, hosts, 'mark_map')
        os.makedirs(user_path, exist_ok=True)
        md_file = self.write_markdown(data, hosts, file_name)
        html_file = os.path.join(user_path, f"{file_name}.html")
        func_box.Shell(f'npx markmap-cli --no-open "{md_file}" -o "{html_file}"').read()
        return md_file, html_file

    def split_startswith_txt(self, link_limit, start='http'):
        link = str(link_limit).split()
        links = []
        for i in link:
            if i.startswith(start):
                links.append(i)
        return links

    def global_search_for_files(self, file_path, matching: list):
        file_list = []
        for root, dirs, files in os.walk(file_path):
            for file in files:
                for math in matching:
                    if str(math).lower() in str(file).lower():
                        file_list.append(os.path.join(root, file))
        return file_list



class ExcelHandle:

    def __init__(self, ipaddr, is_client=True):
        if type(is_client) is bool and is_client:
            self.template_excel = os.path.join(func_box.base_path, 'docs/template/【Temp】测试要点.xlsx')
        elif not is_client:
            self.template_excel = os.path.join(func_box.base_path, 'docs/template/接口测试用例模板.xlsx')
        elif type(type) is str:
            if os.path.exists(is_client):
                self.template_excel = is_client
            else:
                self.template_excel = os.path.join(func_box.base_path, 'docs/template/【Temp】测试要点.xlsx')
        self.user_path = os.path.join(func_box.base_path, 'private_upload', ipaddr, 'test_case')
        os.makedirs(f'{self.user_path}', exist_ok=True)

    def lpvoid_lpbuffe(self, data_list: list, filename='', decs=''):
        # 加载现有的 Excel 文件
        workbook = load_workbook(self.template_excel)
        # 选择要操作的工作表
        worksheet = workbook['测试要点']
        try:
            decs_sheet = workbook['说明']
            decs_sheet['C2'] = decs
        except:
            print('文档没有说明的sheet')
        # 定义起始行号
        start_row = 4
        # 遍历数据列表
        for row_data in data_list:
            # 写入每一行的数据到指定的单元格范围
            for col_num, value in enumerate(row_data, start=1):
                cell = worksheet.cell(row=start_row, column=col_num)
                try:
                    cell.value = str(value).strip()
                except Exception:
                    print(row_data, value)
                    func_box.通知机器人(error=f'写入excel错误啦\n\n```\n\n{row_data}\n\n{value}\n\n```\n\n')
            # 增加起始行号
            start_row += 1
        # 保存 Excel 文件
        time_stamp = time.strftime("%Y-%m-%d-%H", time.localtime())
        if filename == '': filename = time.strftime("%Y-%m-%d-%H", time.localtime()) + '_temp'
        else: f"{time_stamp}_{filename}"
        test_case_path = f'{os.path.join(self.user_path, filename)}.xlsx'
        workbook.save(test_case_path)
        return test_case_path


def get_docs_content(url, image_processing=False):
    """
    Args: 爬虫程序，通过拿到分享链接提取文档内信息
        url: 文档url
        image_processing: 是否开始OCR
    Returns:
    """
    kdocs = crzay_kingsoft.Kdocs(url)
    utils = Utils()
    json_data, json_dict = kdocs.get_file_content()
    text_values = utils.find_all_text_keys(json_data, filter_type='')
    _all, content, pic_dict, file_dict = utils.statistical_results(text_values, img_proce=image_processing)
    pic_dict_convert = kdocs.get_file_pic_url(pic_dict)
    empty_picture_count = sum(1 for item in _all if 'picture' in item and not item['picture']['caption'])
    return _all, content, empty_picture_count, pic_dict_convert, file_dict


def batch_recognition_images_to_md(img_list, ipaddr):
    """
    Args: 将图片批量识别然后写入md文件
        img_list: 图片地址list
        ipaddr: 用户所属标识
    Returns: [文件list]
    """
    temp_list = []
    for img in img_list:
        if os.path.exists(img):
            img_content, img_result = ocr_tools.Paddle_ocr_select(ipaddr=ipaddr, trust_value=True
                                                                  ).img_def_content(img_path=img)
            temp_file = os.path.join(func_box.users_path, ipaddr, 'ocr_to_md', img_content.splitlines()[0][:20]+'.md')
            with open(temp_file, mode='w') as f:
                f.write(f"{func_box.html_view_blank(temp_file)}\n\n"+img_content)
            temp_list.append(temp_list)
        else:
            print(img, '文件路径不存在')
    return temp_list


def get_kdocs_dir(limit, project_folder, type, ipaddr):
    """
    Args:
        limit: 文档目录路径
        project_folder: 写入的文件
        type: 文件类型, 不过这里没用到
        ipaddr:  文件所属标识
    Returns: [文件列表], 目录内文件信息, 失败信息
    """
    kdocs = crzay_kingsoft.Kdocs(limit)
    task_id, task_info = kdocs.submit_batch_download_tasks()
    link, task_faillist = kdocs.polling_batch_download_tasks(task_id)
    resp = kdocs.wps_file_download(link)
    content = resp.content
    temp_file = os.path.join(project_folder, kdocs.url_dirs_tag + '.zip')
    with open(temp_file, 'wb') as f: f.write(content)
    decompress_directory = os.path.join(project_folder, 'extract', kdocs.url_dirs_tag)
    toolbox.extract_archive(temp_file, decompress_directory)
    file_list = []
    img_list = []
    for f_t in kdocs.docs_old_type:
        _, file_, _ = crazy_utils.get_files_from_everything(decompress_directory, type=f_t, ipaddr=ipaddr)
        file_list += file_
    for i_t in Utils().picture_format:
        _, file_, _ = crazy_utils.get_files_from_everything(decompress_directory, type=i_t, ipaddr=ipaddr)
        file_list += file_
    file_list += batch_recognition_images_to_md(img_list, ipaddr)
    return file_list, task_info, task_faillist


def get_kdocs_files(limit, project_folder, type, ipaddr):
    """
    Args:
        limit: 金山文档分享文件地址
        project_folder: 存储地址
        type: 指定的文件类型
        ipaddr: 用户信息
    Returns: [提取的文件list]
    """
    if type == 'otl':
        _, content, _, pic_dict, _ = get_docs_content(limit)
        name = 'temp.md'
        tag = content.splitlines()[0][:20]
        for i in pic_dict:  # 增加OCR选项
            img_content, img_result = ocr_tools.Paddle_ocr_select(ipaddr=ipaddr, trust_value=True
                                                                  ).img_def_content(img_path=pic_dict[i])
            content = str(content).replace(f"{i}", f"{func_box.html_local_img(img_result)}\n```{img_content}```")
            name = tag + '.md'
            content = content.encode('utf-8')
    elif type or type == '':
        kdocs = crzay_kingsoft.Kdocs(limit)
        link, name = kdocs.document_aggregation_download(file_type=type)
        tag = kdocs.url_share_tag
        if link:
            resp = requests.get(url=link, verify=False)
            content = resp.content
        else:
            return []
    else:
        return []
    if content:
        tag_path = os.path.join(project_folder, tag)
        temp_file = os.path.join(os.path.join(project_folder, tag, name))
        os.makedirs(tag_path, exist_ok=True)
        with open(temp_file, 'wb') as f: f.write(content)
        return [temp_file]


def get_kdocs_from_everything(txt, type='', ipaddr='temp'):
    """
    Args:
        txt: kudos 文件分享码
        type: type=='' 时，将支持所有文件类型
        ipaddr: 用户信息
    Returns:
    """
    link_limit = Utils().split_startswith_txt(link_limit=txt)
    file_manifest = []
    success = ''
    project_folder = os.path.join(func_box.users_path, ipaddr, 'kdocs')
    os.makedirs(project_folder, exist_ok=True)
    if link_limit:
        for limit in link_limit:
            if '/ent/' in limit:
                file_list, info, fail = get_kdocs_dir(limit, project_folder, type, ipaddr)
                file_manifest += file_list
                success += f"{limit}文件信息如下：{info}\n\n 下载任务状况：{fail}\n\n"
            else:
                file_manifest += get_kdocs_files(limit, project_folder, type, ipaddr)
    return success, file_manifest, project_folder

def json_args_return(kwargs, keys: list) -> list:
    """
    Args: 提取插件的调优参数，如果有，则返回取到的值，如果无，则返回False
        kwargs: 一般是plugin_kwargs
        keys: 需要取得key
    Returns: 有key返value，无key返False
    """
    temp = [False for i in range(len(keys))]
    for i in range(len(keys)):
        try:
            temp[i] = json.loads(kwargs['advanced_arg'])[keys[i]]
        except Exception as f:
            try:
                temp[i] = kwargs['parameters_def'][keys[i]]
            except Exception as f:
                temp[i] = False
    return temp

def replace_special_chars(file_name):
    # 除了中文外，该正则表达式匹配任何一个不是数字、字母、下划线、.、空格的字符，避免文件名有问题
    new_name = re.sub(r'[^\u4e00-\u9fa5\d\w\s\.\_]', '', file_name)
    if not new_name:
        new_name = func_box.created_atime()
    return new_name


def long_name_processing(file_name):
    """
    Args:
        file_name: 文件名取材，如果是list，则取下标0，转换为str， 如果是str则取最多20个字符
    Returns: 返回处理过的文件名
    """
    if type(file_name) is list: file_name = file_name[0]
    if len(file_name) > 50:
        if file_name.find('"""') != -1:
            temp = file_name.split('"""')[1].splitlines()
            for i in temp:
                if i:
                    file_name = replace_special_chars(i)
                    break
        else:
            file_name = file_name[:20]
    return file_name


def write_test_cases(gpt_response_collection, inputs_show_user_array, llm_kwargs, chatbot, history, is_client=True):
    """
    Args:
        gpt_response_collection: [输出， 输出]
        inputs_show_user_array: [输出]
        llm_kwargs: 调优参数
        chatbot: 对话组件
        history: 对话历史
        is_client: 是否客户端测试用例
    Returns: None
    """
    gpt_response = gpt_response_collection[1::2]
    chat_file_list = ''
    test_case = []
    file_name = long_name_processing(inputs_show_user_array)
    for k in range(len(gpt_response)):
        gpt_response_split = gpt_response[k].splitlines()[2:]  # 过滤掉表头
        for i in gpt_response_split:
            if i.find('|') != -1:
                test_case.append([func_box.clean_br_string(i) for i in i.split('|')[1:]])
            elif i.find('｜') != -1:
                test_case.append([func_box.clean_br_string(i) for i in i.split('｜')[1:]])
            else:
                func_box.通知机器人(f'脏数据过滤，这个不符合写入测试用例的条件 \n\n```\n\n{i}\n\n```')
    file_path = ExcelHandle(ipaddr=llm_kwargs['ipaddr'], is_client=is_client).lpvoid_lpbuffe(test_case, filename=file_name)
    chat_file_list += f'{file_name}生成结果如下:\t {func_box.html_download_blank(__href=file_path, dir_name=file_path.split("/")[-1])}\n\n'
    chatbot.append(['Done', chat_file_list])
    yield from toolbox.update_ui(chatbot, history)


def split_content_limit(inputs: str, llm_kwargs, chatbot, history) -> list:
    """
    Args:
        inputs: 需要提取拆分的提问信息
        llm_kwargs: 调优参数
        chatbot: 对话组件
        history: 历史记录
    Returns: [拆分1， 拆分2]
    """
    model = llm_kwargs['llm_model']
    all_tokens = bridge_all.model_info[llm_kwargs['llm_model']]['max_token']
    max_token = all_tokens/2 - all_tokens/4  # 考虑到对话+回答会超过tokens,所以/2
    get_token_num = bridge_all.model_info[model]['token_cnt']
    inputs = inputs.split('\n---\n')
    segments = []
    for input_ in inputs:
        if get_token_num(input_) > max_token:
            chatbot.append([None, f'{func_box.html_tag_color(input_[:10])}...对话预计超出tokens限制, 拆分中...'])
            yield from toolbox.update_ui(chatbot, history)
            segments.extend(crazy_utils.breakdown_txt_to_satisfy_token_limit(input_, get_token_num, max_token))
        else:
            segments.append(input_)
    yield from toolbox.update_ui(chatbot, history)
    return segments


def input_output_processing(gpt_response_collection, llm_kwargs, plugin_kwargs, chatbot, history, default_prompt: str = False, all_chat: bool = True):
    """
    Args:
        gpt_response_collection:  多线程GPT的返回结果
        plugin_kwargs: 对话使用的插件参数
        chatbot: 对话组件
        history: 历史对话
        llm_kwargs:  调优参数
        default_prompt: 默认Prompt, 如果为False，则不添加提示词
    Returns: 下次使用？
        inputs_array， inputs_show_user_array
    """
    inputs_array = []
    inputs_show_user_array = []
    kwargs_prompt, prompt_cls = json_args_return(plugin_kwargs, ['prompt', 'prompt_cls'])
    if default_prompt: kwargs_prompt = default_prompt
    chatbot.append([f'接下来使用的Prompt是 {func_box.html_tag_color(kwargs_prompt)} ，'
                     f'你可以保存一个同名的Prompt，或在{func_box.html_tag_color("自定义插件参数")}中指定另一个Prmopt哦～', None])
    time.sleep(1)
    prompt = prompt_generator.SqliteHandle(table=f'prompt_{llm_kwargs["ipaddr"]}').find_prompt_result(kwargs_prompt, prompt_cls)
    for inputs, you_say in zip(gpt_response_collection[1::2], gpt_response_collection[0::2]):
        content_limit = yield from split_content_limit(inputs, llm_kwargs, chatbot, history)
        for limit in content_limit:
            inputs_array.append(prompt.replace('{{{v}}}', limit))
            inputs_show_user_array.append(you_say)
    yield from toolbox.update_ui(chatbot, history)
    if all_chat:
        inputs_show_user_array = inputs_array
    else:
        inputs_show_user_array = default_prompt + ': ' + gpt_response_collection[0::2]
    return inputs_array, inputs_show_user_array


def submit_multithreaded_tasks(inputs_array, inputs_show_user_array, llm_kwargs, chatbot, history, plugin_kwargs):
    """
    Args: 提交多线程任务
        inputs_array: 需要提交给gpt的任务列表
        inputs_show_user_array: 显示在user页面上信息
        llm_kwargs: 调优参数
        chatbot: 对话组件
        history: 历史对话
        plugin_kwargs: 插件调优参数
    Returns:  将对话结果返回[输入, 输出]
    """
    if len(inputs_array) == 1:
        # 折叠输出
        if len(inputs_array[0]) > 200: inputs_show_user = \
            inputs_array[0][:100]+f"\n\n{func_box.html_tag_color('......超过200个字符折叠......')}\n\n"+inputs_array[0][-100:]
        else: inputs_show_user = inputs_array[0]
        gpt_say = yield from crazy_utils.request_gpt_model_in_new_thread_with_ui_alive(
            inputs=inputs_array[0], inputs_show_user=inputs_show_user,
            llm_kwargs=llm_kwargs, chatbot=chatbot, history=[],
            sys_prompt="", refresh_interval=0.1
        )
        gpt_response_collection = [inputs_show_user_array[0], gpt_say]
        history.extend(gpt_response_collection)
    else:
        gpt_response_collection = yield from crazy_utils.request_gpt_model_multi_threads_with_very_awesome_ui_and_high_efficiency(
            inputs_array=inputs_array,
            inputs_show_user_array=inputs_show_user_array,
            llm_kwargs=llm_kwargs,
            chatbot=chatbot,
            history_array=[[""] for _ in range(len(inputs_array))],
            sys_prompt_array=["" for _ in range(len(inputs_array))],
            # max_workers=5,  # OpenAI所允许的最大并行过载
            scroller_max_len=80
        )
        # 是否展示任务结果
        kwargs_is_show,  = json_args_return(plugin_kwargs, ['is_show'])
        if kwargs_is_show:
            for results in list(zip(gpt_response_collection[0::2], gpt_response_collection[1::2])):
                chatbot.append(results)
                history.extend(results)
                yield from toolbox.update_ui(chatbot, history)
    return gpt_response_collection


def transfer_flow_chart(gpt_response_collection, llm_kwargs, chatbot, history):
    """
    Args: 将输出结果写入md，并转换为流程图
        gpt_response_collection: [输入、输出]
        llm_kwargs: 调优参数
        chatbot: 对话组件
        history: 历史对话
    Returns:
        None
    """
    for inputs, you_say in zip(gpt_response_collection[1::2], gpt_response_collection[0::2]):
        chatbot.append([None, f'{long_name_processing(you_say)} 🏃🏻‍正在努力将Markdown转换为流程图~'])
        yield from toolbox.update_ui(chatbot=chatbot, history=history)
        inputs = str(inputs).lstrip('```').rstrip('```')  # 去除头部和尾部的代码块, 避免流程图堆在一块
        md, html = Utils().markdown_to_flow_chart(data=inputs, hosts=llm_kwargs['ipaddr'], file_name=long_name_processing(you_say))
        chatbot.append(("View: " + func_box.html_view_blank(md), f'{func_box.html_iframe_code(html_file=html)}'
                                                               f'tips: 双击空白处可以放大～'
                                                               f'\n\n--- \n\n Download: {func_box.html_download_blank(html)}' 
                                                              '\n\n--- \n\n View: ' + func_box.html_view_blank(html)))
        yield from toolbox.update_ui(chatbot=chatbot, history=history, msg='成功写入文件！')


def result_written_to_markdwon(gpt_response_collection, llm_kwargs, chatbot, history):
    """
    Args: 将输出结果写入md
        gpt_response_collection: [输入、输出]
        llm_kwargs: 调优参数
        chatbot: 对话组件
        history: 历史对话
    Returns:
        None
    """
    inputs_all = ''
    for inputs, you_say in zip(gpt_response_collection[1::2], gpt_response_collection[0::2]):
        inputs_all += inputs
    md = Utils().write_markdown(data=inputs_all, hosts=llm_kwargs['ipaddr'], file_name=long_name_processing(gpt_response_collection[0]))
    chatbot.append((None, f'markdown已写入文件，下次可以直接提交markdown文件，就可以节省tomarkdown的时间啦 {func_box.html_view_blank(md)}'))
    yield from toolbox.update_ui(chatbot=chatbot, history=history, msg='成功写入文件！')


previously_on_plugins = f'如果是本地文件，请点击【🔗】先上传，多个文件请上传压缩包，'\
                  f'如果是网络文件或金山文档链接，{func_box.html_tag_color("请粘贴到输入框, 然后再次点击该插件")}'\
                  f'多个文件{func_box.html_tag_color("请使用换行或空格区分")}'


if __name__ == '__main__':
    print(get_docs_content('https://www.kdocs.cn/l/cnYprFmFqghk'))
