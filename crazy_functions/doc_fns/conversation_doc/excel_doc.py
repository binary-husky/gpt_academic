import re
import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook


class ExcelTableFormatter:
    """聊天记录中Markdown表格转Excel生成器"""

    def __init__(self):
        """初始化Excel文档对象"""
        self.workbook = Workbook()
        self._table_count = 0
        self._current_sheet = None

    def _normalize_table_row(self, row):
        """标准化表格行，处理不同的分隔符情况"""
        row = row.strip()
        if row.startswith('|'):
            row = row[1:]
        if row.endswith('|'):
            row = row[:-1]
        return [cell.strip() for cell in row.split('|')]

    def _is_separator_row(self, row):
        """检查是否是分隔行（由 - 或 : 组成）"""
        clean_row = re.sub(r'[\s|]', '', row)
        return bool(re.match(r'^[-:]+$', clean_row))

    def _extract_tables_from_text(self, text):
        """从文本中提取所有表格内容"""
        if not isinstance(text, str):
            return []

        tables = []
        current_table = []
        is_in_table = False

        for line in text.split('\n'):
            line = line.strip()
            if not line:
                if is_in_table and current_table:
                    if len(current_table) >= 2:
                        tables.append(current_table)
                    current_table = []
                    is_in_table = False
                continue

            if '|' in line:
                if not is_in_table:
                    is_in_table = True
                current_table.append(line)
            else:
                if is_in_table and current_table:
                    if len(current_table) >= 2:
                        tables.append(current_table)
                    current_table = []
                    is_in_table = False

        if is_in_table and current_table and len(current_table) >= 2:
            tables.append(current_table)

        return tables

    def _parse_table(self, table_lines):
        """解析表格内容为结构化数据"""
        try:
            headers = self._normalize_table_row(table_lines[0])

            separator_index = next(
                (i for i, line in enumerate(table_lines) if self._is_separator_row(line)),
                1
            )

            data_rows = []
            for line in table_lines[separator_index + 1:]:
                cells = self._normalize_table_row(line)
                # 确保单元格数量与表头一致
                while len(cells) < len(headers):
                    cells.append('')
                cells = cells[:len(headers)]
                data_rows.append(cells)

            if headers and data_rows:
                return {
                    'headers': headers,
                    'data': data_rows
                }
        except Exception as e:
            print(f"解析表格时发生错误: {str(e)}")

        return None

    def _create_sheet(self, question_num, table_num):
        """创建新的工作表"""
        sheet_name = f'Q{question_num}_T{table_num}'
        if len(sheet_name) > 31:
            sheet_name = f'Table{self._table_count}'

        if sheet_name in self.workbook.sheetnames:
            sheet_name = f'{sheet_name}_{datetime.now().strftime("%H%M%S")}'

        return self.workbook.create_sheet(title=sheet_name)

    def create_document(self, history):
        """
        处理聊天历史中的所有表格并创建Excel文档

        Args:
            history: 聊天历史列表

        Returns:
            Workbook: 处理完成的Excel工作簿对象，如果没有表格则返回None
        """
        has_tables = False

        # 删除默认创建的工作表
        default_sheet = self.workbook['Sheet']
        self.workbook.remove(default_sheet)

        # 遍历所有回答
        for i in range(1, len(history), 2):
            answer = history[i]
            tables = self._extract_tables_from_text(answer)

            for table_lines in tables:
                parsed_table = self._parse_table(table_lines)
                if parsed_table:
                    self._table_count += 1
                    sheet = self._create_sheet(i // 2 + 1, self._table_count)

                    # 写入表头
                    for col, header in enumerate(parsed_table['headers'], 1):
                        sheet.cell(row=1, column=col, value=header)

                    # 写入数据
                    for row_idx, row_data in enumerate(parsed_table['data'], 2):
                        for col_idx, value in enumerate(row_data, 1):
                            sheet.cell(row=row_idx, column=col_idx, value=value)

                    has_tables = True

        return self.workbook if has_tables else None


def save_chat_tables(history, save_dir, base_name):
    """
    保存聊天历史中的表格到Excel文件

    Args:
        history: 聊天历史列表
        save_dir: 保存目录
        base_name: 基础文件名

    Returns:
        list: 保存的文件路径列表
    """
    result_files = []

    try:
        # 创建Excel格式
        excel_formatter = ExcelTableFormatter()
        workbook = excel_formatter.create_document(history)

        if workbook is not None:
            # 确保保存目录存在
            os.makedirs(save_dir, exist_ok=True)

            # 生成Excel文件路径
            excel_file = os.path.join(save_dir, base_name + '.xlsx')

            # 保存Excel文件
            workbook.save(excel_file)
            result_files.append(excel_file)
            print(f"已保存表格到Excel文件: {excel_file}")
    except Exception as e:
        print(f"保存Excel格式失败: {str(e)}")

    return result_files


# 使用示例
if __name__ == "__main__":
    # 示例聊天历史
    history = [
        "问题1",
        """这是第一个表格：
        | A | B | C |
        |---|---|---|
        | 1 | 2 | 3 |""",

        "问题2",
        "这是没有表格的回答",

        "问题3",
        """回答包含多个表格：
        | Name | Age |
        |------|-----|
        | Tom  | 20  |

        第二个表格：
        | X | Y |
        |---|---|
        | 1 | 2 |"""
    ]

    # 保存表格
    save_dir = "output"
    base_name = "chat_tables"
    saved_files = save_chat_tables(history, save_dir, base_name)